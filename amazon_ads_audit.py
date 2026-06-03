#!/usr/bin/env python3
"""
Amazon Ads Bulk Sheet → Audit Report generator.

Reads official Amazon Ads bulk export (.xlsx) and writes a 3-sheet audit workbook.
"""

from __future__ import annotations

import argparse
import re
import sys
import warnings
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Styling constants (match BlackBuck audit report)
# ---------------------------------------------------------------------------
COLOR_TITLE = "1F3864"
COLOR_SECTION = "2E75B6"
COLOR_SECTION_ALT = "2F5597"
COLOR_SUBHEADER = "BDD7EE"
COLOR_ROW_ALT = "EBF3FB"
COLOR_WHITE = "FFFFFF"

FONT_TITLE = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_SUBHEADER = Font(name="Calibri", size=11, bold=True, color="000000")
FONT_DATA = Font(name="Calibri", size=11)
FONT_DATA_BOLD = Font(name="Calibri", size=11, bold=True)

FILL_TITLE = PatternFill("solid", fgColor=COLOR_TITLE)
FILL_SECTION = PatternFill("solid", fgColor=COLOR_SECTION)
FILL_SECTION_ALT = PatternFill("solid", fgColor=COLOR_SECTION_ALT)
FILL_SUBHEADER = PatternFill("solid", fgColor=COLOR_SUBHEADER)
FILL_ROW_ALT = PatternFill("solid", fgColor=COLOR_ROW_ALT)
FILL_WHITE = PatternFill("solid", fgColor=COLOR_WHITE)

FMT_CURRENCY = '\u20b9#,##0.00'
FMT_PCT = "0.00%"  # reserved for true Excel percent cells
FMT_RATIO = "0.000000"  # ACoS, CTR, Conv. Rate, Spend %, Sales % (decimal ratios)
FMT_INT = "#,##0"
FMT_CPC = '\u20b9#,##0.00'

METRIC_HEADERS = [
    "Spend",
    "Spend %",
    "Sales",
    "Sales %",
    "ACoS",
    "Clicks",
    "CPC",
    "Orders",
    "Conv. Rate",
    "Impressions",
    "CTR",
]

SHEET_SP = "Sponsored Products Campaigns"
SHEET_SB = "Sponsored Brands Campaigns"
SHEET_SB_MAG = "SB Multi Ad Group Campaigns"
SHEET_SD = "Sponsored Display Campaigns"

PLACEMENT_MAP = {
    "placement top": "Top Of Search",
    "placement rest of search": "Rest of the Search",
    "placement product page": "Product Pages",
}

AUTO_TARGET_MAP = {
    "close-match": "Close Match",
    "loose-match": "Loose Match",
    "substitutes": "Substitutes",
    "complements": "Complements",
}

POSITIVE_MATCH_TYPES = ("Broad", "Phrase", "Exact")


# ---------------------------------------------------------------------------
# Column detection & data loading
# ---------------------------------------------------------------------------
def _normalize_col(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).strip().lower())


def detect_metric_columns(df: pd.DataFrame) -> dict[str, str]:
    """Map logical metric names to actual bulk-sheet column names."""
    cols = {_normalize_col(c): c for c in df.columns}
    mapping: dict[str, str] = {}

    def pick(*candidates: str) -> str | None:
        for cand in candidates:
            key = _normalize_col(cand)
            if key in cols:
                return cols[key]
        return None

    for base in ("impressions", "clicks", "spend"):
        found = pick(base) or pick(base.title())
        if found:
            mapping[base] = found

    sales = None
    for pat in (
        r"14\s*day\s*total\s*sales",
        r"7\s*day\s*total\s*sales",
        r"^sales$",
        r"sales\s*\(#\)",
    ):
        for norm, orig in cols.items():
            if re.search(pat, norm):
                sales = orig
                break
        if sales:
            break
    if sales:
        mapping["sales"] = sales

    orders = None
    for pat in (
        r"14\s*day\s*total\s*orders",
        r"7\s*day\s*total\s*orders",
        r"^orders$",
        r"orders\s*\(#\)",
    ):
        for norm, orig in cols.items():
            if re.search(pat, norm):
                orders = orig
                break
        if orders:
            break
    if orders:
        mapping["orders"] = orders

    for base, patterns in (
        ("acos", (r"^acos$",)),
        ("cpc", (r"^cpc$",)),
        ("ctr", (r"click.?through", r"^ctr$")),
        ("conversion_rate", (r"conversion\s*rate",)),
    ):
        if base in mapping:
            continue
        for pat in patterns:
            for norm, orig in cols.items():
                if re.search(pat, norm):
                    mapping[base] = orig
                    break
            if base in mapping:
                break

    return mapping


def _numeric_series(df: pd.DataFrame, col: str | None) -> pd.Series:
    if col is None or col not in df.columns:
        return pd.Series(0.0, index=df.index)
    return pd.to_numeric(df[col], errors="coerce").fillna(0)


def load_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        try:
            df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
        except ValueError:
            return pd.DataFrame()
    if df.empty:
        return df
    df.columns = [str(c).strip() for c in df.columns]
    return df


def load_bulk_data(path: Path) -> dict[str, Any]:
    sp = load_sheet(path, SHEET_SP)
    sb = load_sheet(path, SHEET_SB)
    sb_mag = load_sheet(path, SHEET_SB_MAG)
    sd = load_sheet(path, SHEET_SD)

    if sp.empty and sb.empty and sd.empty:
        raise ValueError(
            "No recognized campaign sheets found. Expected at least one of: "
            f"'{SHEET_SP}', '{SHEET_SB}', '{SHEET_SD}'."
        )

    return {
        "sp": sp,
        "sb": pd.concat([sb, sb_mag], ignore_index=True) if not sb.empty or not sb_mag.empty else pd.DataFrame(),
        "sd": sd,
        "sp_cols": detect_metric_columns(sp) if not sp.empty else {},
        "sb_cols": detect_metric_columns(sb) if not sb.empty else detect_metric_columns(sb_mag),
        "sd_cols": detect_metric_columns(sd) if not sd.empty else {},
    }


# ---------------------------------------------------------------------------
# Metrics aggregation
# ---------------------------------------------------------------------------
def _entity_mask(df: pd.DataFrame, *entities: str) -> pd.Series:
    if df.empty or "Entity" not in df.columns:
        return pd.Series(False, index=df.index)
    return df["Entity"].astype(str).str.strip().isin(entities)


def aggregate_rows(
    df: pd.DataFrame,
    mask: pd.Series,
    colmap: dict[str, str],
) -> dict[str, float]:
    sub = df.loc[mask] if mask is not None else df
    if sub.empty:
        return _empty_metrics()

    impr = _numeric_series(sub, colmap.get("impressions")).sum()
    clicks = _numeric_series(sub, colmap.get("clicks")).sum()
    spend = _numeric_series(sub, colmap.get("spend")).sum()
    sales = _numeric_series(sub, colmap.get("sales")).sum()
    orders = _numeric_series(sub, colmap.get("orders")).sum()

    return _metrics_from_totals(impr, clicks, spend, sales, orders)


def _empty_metrics() -> dict[str, float]:
    return {
        "spend": 0.0,
        "sales": 0.0,
        "orders": 0.0,
        "clicks": 0.0,
        "impressions": 0.0,
        "acos": np.nan,
        "cpc": np.nan,
        "ctr": 0.0,
        "conv_rate": 0.0,
    }


def _metrics_from_totals(
    impr: float, clicks: float, spend: float, sales: float, orders: float
) -> dict[str, float]:
    acos = spend / sales if sales > 0 else np.nan
    cpc = spend / clicks if clicks > 0 else np.nan
    ctr = clicks / impr if impr > 0 else 0.0
    conv = orders / clicks if clicks > 0 else 0.0
    return {
        "spend": float(spend),
        "sales": float(sales),
        "orders": float(orders),
        "clicks": float(clicks),
        "impressions": float(impr),
        "acos": acos,
        "cpc": cpc,
        "ctr": ctr,
        "conv_rate": conv,
    }


def add_share_metrics(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    total_spend = sum(r["metrics"]["spend"] for r in rows)
    total_sales = sum(r["metrics"]["sales"] for r in rows)
    for r in rows:
        m = r["metrics"]
        m["spend_pct"] = m["spend"] / total_spend if total_spend > 0 else 0.0
        m["sales_pct"] = m["sales"] / total_sales if total_sales > 0 else 0.0
    return rows


def _campaign_ids(df: pd.DataFrame, targeting: str) -> set:
    if df.empty or "Entity" not in df.columns:
        return set()
    camp = df[_entity_mask(df, "Campaign")]
    if "Targeting Type" not in camp.columns:
        return set()
    ids = camp[camp["Targeting Type"].astype(str).str.lower() == targeting.lower()]
    if "Campaign ID" not in ids.columns:
        return set()
    return set(ids["Campaign ID"].dropna().unique())


def _video_campaign_ids(sb: pd.DataFrame) -> set:
    if sb.empty:
        return set()
    camp = sb[_entity_mask(sb, "Campaign", "Draft Campaign")]
    if "Ad Format" not in camp.columns:
        return set()
    vid = camp[camp["Ad Format"].astype(str).str.lower() == "video"]
    if "Campaign ID" not in vid.columns:
        return set()
    return set(vid["Campaign ID"].dropna().unique())


def _map_placement(val: Any) -> str | None:
    if pd.isna(val):
        return None
    key = str(val).strip().lower()
    return PLACEMENT_MAP.get(key)


def _map_auto_target(expr: Any) -> str | None:
    if pd.isna(expr):
        return None
    key = str(expr).strip().lower()
    return AUTO_TARGET_MAP.get(key, str(expr).title())


def build_opd_sections(data: dict[str, Any]) -> list[dict[str, Any]]:
    """Return ordered list of OPD sections for writing."""
    sp, sb, sd = data["sp"], data["sb"], data["sd"]
    sp_c, sb_c, sd_c = data["sp_cols"], data["sb_cols"], data["sd_cols"]
    sections: list[dict[str, Any]] = []

    # --- Account level ---
    account_rows = []
    sp_m = aggregate_rows(sp, _entity_mask(sp, "Campaign"), sp_c)
    account_rows.append({"label": "Sponsored Product", "metrics": sp_m})

    vid_ids = _video_campaign_ids(sb)
    if not sb.empty and "Campaign ID" in sb.columns:
        sb_non_vid = sb[~sb["Campaign ID"].isin(vid_ids)] if vid_ids else sb
        sb_vid = sb[sb["Campaign ID"].isin(vid_ids)] if vid_ids else sb.iloc[0:0]
    else:
        sb_non_vid, sb_vid = sb, sb.iloc[0:0]

    sb_m = _sb_account_metrics(sb_non_vid, sb_c)
    account_rows.append({"label": "Sponsored Brand", "metrics": sb_m})

    sbv_m = _sb_account_metrics(sb_vid, sb_c) if not sb_vid.empty else _empty_metrics()
    account_rows.append({"label": "Sponsored Brand Video", "metrics": sbv_m})

    sd_cpc_m = aggregate_rows(
        sd,
        _entity_mask(sd, "Campaign") & (sd["Cost Type"].astype(str).str.lower() == "cpc")
        if not sd.empty and "Cost Type" in sd.columns
        else pd.Series(False, index=sd.index),
        sd_c,
    )
    sd_vcpm_m = aggregate_rows(
        sd,
        _entity_mask(sd, "Campaign") & (sd["Cost Type"].astype(str).str.lower() == "vcpm")
        if not sd.empty and "Cost Type" in sd.columns
        else pd.Series(False, index=sd.index),
        sd_c,
    )
    account_rows.append({"label": "Sponsored Display (CPC)", "metrics": sd_cpc_m})
    account_rows.append({"label": "Sponsored Display (vCPM)", "metrics": sd_vcpm_m})
    account_rows = add_share_metrics(account_rows)
    sections.append(
        {
            "title": "Overall Account Ad Type Wise Metrics For Last 30 days",
            "section": "ACCOUNT LEVEL",
            "subsection": None,
            "header_label": "Ad Product / Type",
            "rows": account_rows,
            "total": True,
        }
    )

    # --- SP Auto vs Manual ---
    auto_ids = _campaign_ids(sp, "auto")
    manual_ids = _campaign_ids(sp, "manual")
    sp_auto = aggregate_rows(
        sp, _entity_mask(sp, "Campaign") & sp["Campaign ID"].isin(auto_ids), sp_c
    )
    sp_manual = aggregate_rows(
        sp, _entity_mask(sp, "Campaign") & sp["Campaign ID"].isin(manual_ids), sp_c
    )
    sp_breakdown = add_share_metrics(
        [
            {"label": "Automatic", "metrics": sp_auto},
            {"label": "Manual", "metrics": sp_manual},
        ]
    )
    sections.append(
        {
            "title": None,
            "section": "SPONSORED PRODUCT (Overall + Placements)",
            "subsection": "Sponsored Product",
            "header_label": "Ad Product / Type",
            "rows": sp_breakdown,
            "total": True,
        }
    )

    # --- Placements ---
    place_rows = []
    if not sp.empty and "Placement" in sp.columns:
        ba = sp[_entity_mask(sp, "Bidding Adjustment")].copy()
        ba["_place"] = ba["Placement"].apply(_map_placement)
        for label in ("Top Of Search", "Rest of the Search", "Product Pages"):
            mask = ba["_place"] == label
            place_rows.append(
                {"label": label, "metrics": aggregate_rows(ba, mask, sp_c)}
            )
    place_rows = add_share_metrics(place_rows)
    sections.append(
        {
            "title": None,
            "section": None,
            "subsection": "Placement",
            "header_label": "Placement",
            "rows": place_rows,
            "total": True,
        }
    )

    # --- Auto targets ---
    auto_tgt_rows = []
    if auto_ids:
        pt = sp[
            _entity_mask(sp, "Product Targeting") & sp["Campaign ID"].isin(auto_ids)
        ].copy()
        expr_col = "Product Targeting Expression"
        if expr_col in pt.columns:
            pt["_tgt"] = pt[expr_col].apply(_map_auto_target)
            for label in ("Close Match", "Loose Match", "Substitutes", "Complements"):
                mask = pt["_tgt"] == label
                auto_tgt_rows.append(
                    {"label": label, "metrics": aggregate_rows(pt, mask, sp_c)}
                )
    auto_tgt_rows = add_share_metrics(auto_tgt_rows)
    sections.append(
        {
            "title": "SPONSORED PRODUCT (On Targets Level)",
            "section": None,
            "subsection": "Automatic Targets",
            "header_label": "Targeting",
            "rows": auto_tgt_rows,
            "total": True,
        }
    )

    # --- Manual targets ---
    manual_tgt = []
    if manual_ids:
        kw_m = aggregate_rows(
            sp,
            _entity_mask(sp, "Keyword")
            & sp["Campaign ID"].isin(manual_ids)
            & sp["Match Type"].isin(POSITIVE_MATCH_TYPES),
            sp_c,
        )
        pt_m = aggregate_rows(
            sp,
            _entity_mask(sp, "Product Targeting") & sp["Campaign ID"].isin(manual_ids),
            sp_c,
        )
        manual_tgt = add_share_metrics(
            [
                {"label": "Keyword", "metrics": kw_m},
                {"label": "Product", "metrics": pt_m},
            ]
        )
    sections.append(
        {
            "title": None,
            "section": None,
            "subsection": "Manual Targets",
            "header_label": "Targeting",
            "rows": manual_tgt,
            "total": True,
        }
    )

    # --- Match types ---
    match_rows = []
    if manual_ids:
        kw = sp[
            _entity_mask(sp, "Keyword")
            & sp["Campaign ID"].isin(manual_ids)
            & sp["Match Type"].isin(POSITIVE_MATCH_TYPES)
        ]
        for mt in POSITIVE_MATCH_TYPES:
            mask = kw["Match Type"] == mt
            match_rows.append(
                {"label": mt, "metrics": aggregate_rows(kw, mask, sp_c)}
            )
    match_rows = add_share_metrics(match_rows)
    sections.append(
        {
            "title": None,
            "section": None,
            "subsection": "Keyword Targets (Match Type)",
            "header_label": "Match Type",
            "rows": match_rows,
            "total": True,
        }
    )

    # --- SB keyword vs product ---
    sb_rows = _sb_targeting_breakdown(sb, sb_c)
    sections.append(
        {
            "title": "SPONSORED BRAND (Detailed)",
            "section": None,
            "subsection": "Sponsored Brand (Keywords)",
            "header_label": "Type",
            "rows": sb_rows,
            "total": True,
        }
    )

    # --- SD contextual vs audience ---
    sd_rows = _sd_targeting_breakdown(sd, sd_c)
    sections.append(
        {
            "title": "SPONSORED DISPLAY",
            "section": None,
            "subsection": "Sponsored Display (Targeting)",
            "header_label": "Type",
            "rows": sd_rows,
            "total": True,
        }
    )

    return sections


def _sb_account_metrics(sb: pd.DataFrame, colmap: dict[str, str]) -> dict[str, float]:
    if sb.empty:
        return _empty_metrics()
    m = aggregate_rows(sb, _entity_mask(sb, "Campaign", "Draft Campaign"), colmap)
    if m["spend"] > 0 or m["impressions"] > 0:
        return m
    # Fallback: leaf entities when campaign rows lack metrics
    mask = _entity_mask(sb, "Keyword", "Product Targeting", "Product Collection Ad", "Video Ad", "Brand Video Ad")
    return aggregate_rows(sb, mask, colmap)


def _sb_targeting_breakdown(sb: pd.DataFrame, colmap: dict[str, str]) -> list[dict[str, Any]]:
    if sb.empty:
        return add_share_metrics(
            [
                {"label": "Keyword", "metrics": _empty_metrics()},
                {"label": "Product Targeting", "metrics": _empty_metrics()},
            ]
        )
    kw_m = aggregate_rows(sb, _entity_mask(sb, "Keyword"), colmap)
    pt_mask = _entity_mask(sb, "Product Targeting")
    if not pt_mask.any():
        pt_mask = _entity_mask(sb, "Product Collection Ad", "Video Ad")
    pt_m = aggregate_rows(sb, pt_mask, colmap)
    return add_share_metrics(
        [
            {"label": "Keyword", "metrics": kw_m},
            {"label": "Product Targeting", "metrics": pt_m},
        ]
    )


def _sd_targeting_breakdown(sd: pd.DataFrame, colmap: dict[str, str]) -> list[dict[str, Any]]:
    if sd.empty:
        return add_share_metrics(
            [
                {"label": "Contextual", "metrics": _empty_metrics()},
                {"label": "Audience", "metrics": _empty_metrics()},
            ]
        )
    ctx_m = aggregate_rows(sd, _entity_mask(sd, "Contextual Targeting"), colmap)
    aud_m = aggregate_rows(sd, _entity_mask(sd, "Audience Targeting"), colmap)
    if ctx_m["spend"] == 0 and aud_m["spend"] == 0:
        # Fallback: split by tactic codes when targeting entities lack metrics
        if "Tactic" in sd.columns:
            ctx_m = aggregate_rows(
                sd, sd["Tactic"].astype(str).str.upper() == "T00020", colmap
            )
            aud_m = aggregate_rows(
                sd, sd["Tactic"].astype(str).str.upper() == "T00030", colmap
            )
    return add_share_metrics(
        [
            {"label": "Contextual", "metrics": ctx_m},
            {"label": "Audience", "metrics": aud_m},
        ]
    )


def build_asin_report(data: dict[str, Any]) -> pd.DataFrame:
    """Per ASIN/SKU aggregated metrics, sorted by spend descending."""
    frames = []
    sp, sd = data["sp"], data["sd"]
    sp_c, sd_c = data["sp_cols"], data["sd_cols"]

    for df, colmap, label in ((sp, sp_c, "SP"), (sd, sd_c, "SD")):
        if df.empty:
            continue
        pa = df[_entity_mask(df, "Product Ad")].copy()
        if pa.empty:
            continue
        asin_col = None
        for c in ("ASIN (Informational only)", "ASIN", "Advertised ASIN"):
            if c in pa.columns:
                asin_col = c
                break
        sku_col = "SKU" if "SKU" in pa.columns else None
        if asin_col is None and sku_col is None:
            continue

        pa["_asin"] = pa[asin_col].astype(str) if asin_col else ""
        pa["_sku"] = pa[sku_col].astype(str) if sku_col else ""
        pa["_key"] = pa["_asin"].where(
            pa["_asin"].notna() & (pa["_asin"] != "nan") & (pa["_asin"] != ""),
            pa["_sku"],
        )

        for key, grp in pa.groupby("_key", dropna=True):
            if not key or key == "nan":
                continue
            m = aggregate_rows(grp, pd.Series(True, index=grp.index), colmap)
            row_asin = grp["_asin"].iloc[0] if asin_col else key
            row_sku = grp["_sku"].iloc[0] if sku_col else ""
            if str(row_asin) in ("nan", ""):
                row_asin = key
            frames.append(
                {
                    "ASINs": row_asin,
                    "SKU": row_sku,
                    "Impressions": m["impressions"],
                    "Clicks": m["clicks"],
                    "Spend": m["spend"],
                    "Sales": m["sales"],
                    "Orders": m["orders"],
                    "ACoS": m["acos"],
                }
            )

    if not frames:
        return pd.DataFrame(
            columns=["ASINs", "SKU", "Impressions", "Clicks", "Spend", "Sales", "Orders", "ACoS"]
        )

    out = pd.DataFrame(frames)
    out = out.sort_values("Spend", ascending=False).reset_index(drop=True)
    return out


def build_keywords_report(data: dict[str, Any]) -> dict[str, pd.DataFrame]:
    """Exact / Phrase / Broad keyword tables from SP campaigns."""
    sp = data["sp"]
    sp_c = data["sp_cols"]
    result = {}
    if sp.empty or "Keyword Text" not in sp.columns:
        for mt in POSITIVE_MATCH_TYPES:
            result[mt] = pd.DataFrame(
                columns=["Keywords", "Impressions", "Clicks", "Spend", "Sales", "Orders", "CPC", "ACoS"]
            )
        return result

    manual_ids = _campaign_ids(sp, "manual")
    kw = sp[
        _entity_mask(sp, "Keyword")
        & sp["Match Type"].isin(POSITIVE_MATCH_TYPES)
        & (
            sp["Campaign ID"].isin(manual_ids)
            if manual_ids
            else pd.Series(True, index=sp.index)
        )
    ].copy()

    text_col = "Keyword Text"
    for mt in POSITIVE_MATCH_TYPES:
        sub = kw[kw["Match Type"] == mt].copy()
        if sub.empty:
            result[mt] = pd.DataFrame(
                columns=["Keywords", "Impressions", "Clicks", "Spend", "Sales", "Orders", "CPC", "ACoS"]
            )
            continue
        rows = []
        for text, grp in sub.groupby(text_col, dropna=True):
            if pd.isna(text) or str(text).strip() == "":
                continue
            m = aggregate_rows(grp, pd.Series(True, index=grp.index), sp_c)
            rows.append(
                {
                    "Keywords": str(text).strip(),
                    "Impressions": m["impressions"],
                    "Clicks": m["clicks"],
                    "Spend": m["spend"],
                    "Sales": m["sales"],
                    "Orders": m["orders"],
                    "CPC": m["cpc"],
                    "ACoS": m["acos"],
                }
            )
        df_mt = pd.DataFrame(rows)
        if not df_mt.empty:
            df_mt = df_mt.sort_values("Spend", ascending=False).reset_index(drop=True)
        result[mt] = df_mt

    return result


# ---------------------------------------------------------------------------
# Excel writing & styling
# ---------------------------------------------------------------------------
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _write_metric_row(
    ws,
    row: int,
    label: str,
    m: dict[str, float],
    *,
    bold: bool = False,
    alt: bool = False,
) -> None:
    fill = FILL_ROW_ALT if alt else FILL_WHITE
    font = FONT_DATA_BOLD if bold else FONT_DATA
    ws.cell(row, 1, label).font = font
    ws.cell(row, 1).fill = fill

    values = [
        m.get("spend", 0),
        m.get("spend_pct", np.nan),
        m.get("sales", 0),
        m.get("sales_pct", np.nan),
        m.get("acos", np.nan),
        m.get("clicks", 0),
        m.get("cpc", np.nan),
        m.get("orders", 0),
        m.get("conv_rate", 0),
        m.get("impressions", 0),
        m.get("ctr", 0),
    ]
    fmts = [
        FMT_CURRENCY,
        FMT_RATIO,
        FMT_CURRENCY,
        FMT_RATIO,
        FMT_RATIO,
        FMT_INT,
        FMT_CPC,
        FMT_INT,
        FMT_RATIO,
        FMT_INT,
        FMT_RATIO,
    ]
    for col, (val, fmt) in enumerate(zip(values, fmts), start=2):
        cell = ws.cell(row, col, val if not (isinstance(val, float) and np.isnan(val)) else None)
        cell.font = font
        cell.fill = fill
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right")
        cell.border = THIN_BORDER
    ws.cell(row, 1).border = THIN_BORDER


def _sum_metrics(rows: list[dict[str, Any]]) -> dict[str, float]:
    impr = sum(r["metrics"]["impressions"] for r in rows)
    clicks = sum(r["metrics"]["clicks"] for r in rows)
    spend = sum(r["metrics"]["spend"] for r in rows)
    sales = sum(r["metrics"]["sales"] for r in rows)
    orders = sum(r["metrics"]["orders"] for r in rows)
    m = _metrics_from_totals(impr, clicks, spend, sales, orders)
    m["spend_pct"] = 1.0
    m["sales_pct"] = 1.0
    return m


def write_opd_sheet(wb: Workbook, sections: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("OPD", 0)
    ws.sheet_view.showGridLines = True
    row = 1

    for sec_idx, sec in enumerate(sections):
        if sec.get("title"):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            c = ws.cell(row, 1, sec["title"])
            c.font = FONT_TITLE
            c.fill = FILL_TITLE
            c.alignment = Alignment(horizontal="left", vertical="center")
            row += 2

        if sec.get("section"):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            c = ws.cell(row, 1, sec["section"])
            c.font = FONT_SECTION
            c.fill = FILL_SECTION
            row += 1

        if sec.get("subsection"):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            fill = FILL_SECTION_ALT if sec["subsection"] in (
                "Sponsored Product",
                "Placement",
            ) else FILL_SECTION
            c = ws.cell(row, 1, sec["subsection"])
            c.font = FONT_SECTION
            c.fill = fill
            row += 1

        # Header row
        hdr_label = sec.get("header_label", "Ad Product / Type")
        ws.cell(row, 1, hdr_label).font = FONT_SUBHEADER
        ws.cell(row, 1).fill = FILL_SUBHEADER
        for col, h in enumerate(METRIC_HEADERS, start=2):
            cell = ws.cell(row, col, h)
            cell.font = FONT_SUBHEADER
            cell.fill = FILL_SUBHEADER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER
        ws.cell(row, 1).border = THIN_BORDER
        row += 1

        data_rows = sec.get("rows", [])
        for i, dr in enumerate(data_rows):
            _write_metric_row(ws, row, dr["label"], dr["metrics"], alt=(i % 2 == 1))
            row += 1

        if sec.get("total") and data_rows:
            total_m = _sum_metrics(data_rows)
            _write_metric_row(ws, row, "Total", total_m, bold=True)
            row += 1

        row += 1  # blank separator

    ws.freeze_panes = "A3"
    _autosize_columns(ws, max_col=12)


def write_asin_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("ASIN wise report.")
    headers = ["ASINs", "SKU", "Impressions", "Clicks", "Spend", "Sales", "Orders", "ACoS"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(1, col, h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = FILL_TITLE
        cell.alignment = Alignment(horizontal="center")

    for i, rec in df.iterrows():
        r = int(i) + 2
        alt = int(i) % 2 == 1
        fill = FILL_ROW_ALT if alt else FILL_WHITE
        vals = [
            rec.get("ASINs", ""),
            rec.get("SKU", ""),
            rec.get("Impressions", 0),
            rec.get("Clicks", 0),
            rec.get("Spend", 0),
            rec.get("Sales", 0),
            rec.get("Orders", 0),
            rec.get("ACoS", np.nan),
        ]
        fmts = [None, None, FMT_INT, FMT_INT, FMT_CURRENCY, FMT_CURRENCY, FMT_INT, FMT_RATIO]
        for col, (val, fmt) in enumerate(zip(vals, fmts), start=1):
            if isinstance(val, float) and np.isnan(val):
                val = None
            cell = ws.cell(r, col, val)
            cell.font = FONT_DATA
            cell.fill = fill
            if fmt:
                cell.number_format = fmt
            if col > 2:
                cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A2"
    _autosize_columns(ws, max_col=8)


def write_keywords_sheet(wb: Workbook, kw_data: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("Keywords")
    # Layout: Exact cols 1-8, gap col 9, Phrase 10-17, gap 18, Broad 19-26
    blocks = [
        ("Exact", 1),
        ("Phrase", 10),
        ("Broad", 19),
    ]
    sub_headers = ["Keywords", "Impressions", "Clicks", "Spend", "Sales", "Orders", "CPC", "ACoS"]
    fmts = [None, FMT_INT, FMT_INT, FMT_CURRENCY, FMT_CURRENCY, FMT_INT, FMT_CPC, FMT_RATIO]

    for match_name, start_col in blocks:
        ws.merge_cells(
            start_row=1,
            start_column=start_col,
            end_row=1,
            end_column=start_col + 7,
        )
        c = ws.cell(1, start_col, match_name)
        c.font = FONT_SECTION
        c.fill = FILL_SECTION
        c.alignment = Alignment(horizontal="center")
        for j, h in enumerate(sub_headers):
            cell = ws.cell(2, start_col + j, h)
            cell.font = FONT_SUBHEADER
            cell.fill = FILL_SUBHEADER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    max_len = max(
        (len(kw_data.get(mt, pd.DataFrame())) for mt in POSITIVE_MATCH_TYPES),
        default=0,
    )
    for i in range(max_len):
        r = i + 3
        alt = i % 2 == 1
        fill = FILL_ROW_ALT if alt else FILL_WHITE
        for match_name, start_col in blocks:
            df = kw_data.get(match_name, pd.DataFrame())
            if i < len(df):
                rec = df.iloc[i]
                vals = [
                    rec.get("Keywords", ""),
                    rec.get("Impressions", 0),
                    rec.get("Clicks", 0),
                    rec.get("Spend", 0),
                    rec.get("Sales", 0),
                    rec.get("Orders", 0),
                    rec.get("CPC", np.nan),
                    rec.get("ACoS", np.nan),
                ]
            else:
                vals = [""] + [None] * 7
            for j, (val, fmt) in enumerate(zip(vals, fmts)):
                if isinstance(val, float) and np.isnan(val):
                    val = None
                cell = ws.cell(r, start_col + j, val)
                cell.font = FONT_DATA
                cell.fill = fill
                if fmt and val is not None:
                    cell.number_format = fmt
                if j > 0 and val is not None:
                    cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A3"
    for start_col in (1, 10, 19):
        ws.column_dimensions[get_column_letter(start_col)].width = 36
        for off in range(1, 8):
            ws.column_dimensions[get_column_letter(start_col + off)].width = 14


def _autosize_columns(ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 12
        for row in range(1, min(ws.max_row, 80) + 1):
            val = ws.cell(row, col).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)) + 2, 40))
        ws.column_dimensions[letter].width = max_len


def build_audit_workbook(data: dict[str, Any]) -> Workbook:
    """Build the 3-sheet audit workbook from parsed bulk data."""
    sections = build_opd_sections(data)
    asin_df = build_asin_report(data)
    kw_data = build_keywords_report(data)

    wb = Workbook()
    wb.remove(wb.active)
    write_opd_sheet(wb, sections)
    write_asin_sheet(wb, asin_df)
    write_keywords_sheet(wb, kw_data)
    return wb


def generate_audit_report(input_path: Path, output_path: Path) -> None:
    data = load_bulk_data(input_path)
    wb = build_audit_workbook(data)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def generate_audit_report_bytes(
    file_bytes: bytes,
    original_filename: str = "bulk_sheet.xlsx",
) -> tuple[bytes, str]:
    """
    Generate audit report from uploaded file bytes.

    Returns (xlsx_bytes, suggested_download_filename).
    """
    import tempfile
    from io import BytesIO

    suffix = Path(original_filename).suffix.lower()
    if suffix not in (".xlsx", ".xlsm"):
        raise ValueError("Upload must be an Excel file (.xlsx or .xlsm).")

    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = Path(tmp.name)

    try:
        data = load_bulk_data(tmp_path)
        buffer = BytesIO()
        build_audit_workbook(data).save(buffer)
        out_name = f"{Path(original_filename).stem}_Audit_Report.xlsx"
        return buffer.getvalue(), out_name
    finally:
        tmp_path.unlink(missing_ok=True)


def default_output_path(input_path: Path) -> Path:
    stem = input_path.stem
    return input_path.parent / f"{stem}_Audit_Report.xlsx"


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate an Amazon Ads audit report from a bulk sheet (.xlsx)."
    )
    parser.add_argument(
        "bulk_sheet",
        type=Path,
        help="Path to the Amazon Ads bulk export .xlsx file",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path (default: <input>_Audit_Report.xlsx)",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    input_path = args.bulk_sheet.expanduser().resolve()

    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}", file=sys.stderr)
        return 1
    if input_path.suffix.lower() not in (".xlsx", ".xlsm"):
        print("Error: Input must be an Excel file (.xlsx or .xlsm).", file=sys.stderr)
        return 1

    output_path = (
        args.output.expanduser().resolve()
        if args.output
        else default_output_path(input_path)
    )

    try:
        print(f"Reading bulk sheet: {input_path}")
        generate_audit_report(input_path, output_path)
        print(f"Audit report saved: {output_path}")
        return 0
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
